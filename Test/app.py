import os
import json
import re
import sqlite3
from datetime import datetime
from flask import Flask, render_template, request, jsonify, redirect, url_for, g
from werkzeug.utils import secure_filename
import openpyxl
import requests

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['DATABASE'] = os.path.join(os.path.dirname(__file__), 'fishing.db')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- Database ---
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS lakes (
            id TEXT PRIMARY KEY,
            name TEXT,
            county TEXT,
            nearest_town TEXT,
            area REAL,
            max_depth REAL,
            mean_depth REAL,
            littoral_area REAL,
            shore_length REAL,
            latitude REAL,
            longitude REAL,
            fish_species TEXT,
            metadata_json TEXT,
            last_updated TEXT
        );

        CREATE TABLE IF NOT EXISTS survey_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lake_id TEXT NOT NULL,
            lake_name TEXT,
            survey_year INTEGER,
            survey_date TEXT,
            survey_type TEXT DEFAULT 'Standard Survey',
            species TEXT,
            gear TEXT,
            cpue REAL,
            normal_range_cpue TEXT,
            avg_weight REAL,
            normal_range_weight TEXT,
            count INTEGER,
            source_sheet TEXT,
            uploaded_at TEXT,
            FOREIGN KEY (lake_id) REFERENCES lakes(id)
        );

        CREATE INDEX IF NOT EXISTS idx_survey_lake ON survey_data(lake_id);
        CREATE INDEX IF NOT EXISTS idx_survey_species ON survey_data(species);
        CREATE INDEX IF NOT EXISTS idx_survey_year ON survey_data(survey_year);
        CREATE INDEX IF NOT EXISTS idx_survey_type ON survey_data(survey_type);
    ''')
    # Migrate existing databases: add new columns if they don't exist
    try:
        db.execute('ALTER TABLE survey_data ADD COLUMN survey_date TEXT')
    except Exception:
        pass
    try:
        db.execute("ALTER TABLE survey_data ADD COLUMN survey_type TEXT DEFAULT 'Standard Survey'")
    except Exception:
        pass
    db.commit()

with app.app_context():
    init_db()


# --- CPUE Excel Parser ---
def parse_cpue_excel(filepath):
    """Parse the MN DNR CPUE Excel format with multi-lake blocks per sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        i = 0
        while i < len(rows):
            row = rows[i]
            # Detect lake header: col0=lake_name, col1=year (int), col2="ID", col3=lake_id
            if (row[0] and row[1] and row[2] and row[3]
                    and str(row[2]).strip().upper() == 'ID'
                    and _is_year_or_str(row[1])):
                lake_name = str(row[0]).strip()
                survey_year = int(row[1]) if _is_numeric(row[1]) else None
                lake_id = str(row[3]).strip()

                # Skip the column-header row (Species, Gear, CPUE, ...)
                i += 1
                if i < len(rows):
                    i += 1  # skip header

                # Read data rows until blank
                while i < len(rows):
                    drow = rows[i]
                    if not drow[0] or str(drow[0]).strip() == '':
                        i += 1
                        break
                    species = str(drow[0]).strip()
                    gear = str(drow[1]).strip() if drow[1] else ''
                    cpue = _to_float(drow[2])
                    normal_range_cpue = str(drow[3]).strip() if drow[3] else None
                    avg_weight = _to_float(drow[4])
                    normal_range_weight = str(drow[5]).strip() if drow[5] else None
                    count = _to_int(drow[6])

                    if normal_range_cpue == 'nan' or normal_range_cpue == 'None':
                        normal_range_cpue = None
                    if normal_range_weight == 'nan' or normal_range_weight == 'None':
                        normal_range_weight = None

                    records.append({
                        'lake_id': lake_id,
                        'lake_name': lake_name,
                        'survey_year': survey_year,
                        'species': species,
                        'gear': gear,
                        'cpue': cpue,
                        'normal_range_cpue': normal_range_cpue,
                        'avg_weight': avg_weight,
                        'normal_range_weight': normal_range_weight,
                        'count': count,
                        'source_sheet': sheet_name,
                    })
                    i += 1
            else:
                i += 1

    return records

def _is_year_or_str(v):
    if isinstance(v, (int, float)):
        return 1900 <= v <= 2100
    if isinstance(v, str):
        try:
            return 1900 <= int(v) <= 2100
        except ValueError:
            return False
    return False

def _is_numeric(v):
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False

def _to_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def _to_int(v):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


# --- DNR API ---
DNR_API_URL = 'http://services.dnr.state.mn.us/api/lakefinder/by_id/v1/'

def fetch_lake_metadata(lake_id):
    """Fetch metadata from MN DNR LakeFinder API."""
    try:
        resp = requests.get(DNR_API_URL, params={'id': lake_id}, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data.get('status') == 'OK' and data.get('results'):
            return data['results'][0]
    except Exception as e:
        print(f"DNR API error for {lake_id}: {e}")
    return None

def save_lake_metadata(db, lake_id, meta):
    """Upsert lake metadata into the database."""
    point = meta.get('point', {}).get('epsg:4326', [None, None])
    lng, lat = point[0], point[1]
    morph = meta.get('morphology', {})
    species_list = meta.get('fishSpecies', [])
    fish_species = species_list[0] if species_list else ''

    db.execute('''
        INSERT INTO lakes (id, name, county, nearest_town, area, max_depth, mean_depth,
                           littoral_area, shore_length, latitude, longitude, fish_species,
                           metadata_json, last_updated)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(id) DO UPDATE SET
            name=excluded.name, county=excluded.county, nearest_town=excluded.nearest_town,
            area=excluded.area, max_depth=excluded.max_depth, mean_depth=excluded.mean_depth,
            littoral_area=excluded.littoral_area, shore_length=excluded.shore_length,
            latitude=excluded.latitude, longitude=excluded.longitude,
            fish_species=excluded.fish_species, metadata_json=excluded.metadata_json,
            last_updated=excluded.last_updated
    ''', (
        lake_id, meta.get('name'), meta.get('county'), meta.get('nearest_town'),
        morph.get('area'), morph.get('max_depth'), morph.get('mean_depth'),
        morph.get('littoral_area'), morph.get('shore_length'),
        lat, lng, fish_species,
        json.dumps(meta), datetime.now().isoformat()
    ))
    db.commit()


# --- DNR Lake Survey JSON API ---
DNR_SURVEY_API_URL = 'https://maps.dnr.state.mn.us/cgi-bin/lakefinder/detail.cgi'

# Species abbreviation -> common name mapping
# Users can extend this via the /api/species_map endpoint
SPECIES_MAP = {
  "WAE": "walleye", "NOP": "northern pike", "SMB": "smallmouth bass",
  "LMB": "largemouth bass", "BLC": "black crappie", "WHC": "white crappie",
  "BLG": "bluegill", "SUN": "sunfish", "YEP": "yellow perch",
  "SAU": "sauger", "HSA": "hybrid saugeye", "MUS": "muskellunge",
  "TGM": "tiger muskellunge", "BKT": "brook trout", "BNT": "brown trout",
  "RBT": "rainbow trout", "LAT": "lake trout", "CIS": "cisco (tullibee)",
  "WHI": "whitefish", "CCF": "channel catfish", "FDC": "flathead catfish",
  "BLB": "black bullhead", "YBB": "yellow bullhead", "YEB": "yellow bullhead",
  "BRB": "brown bullhead", "FRD": "freshwater drum", "STN": "lake sturgeon",
  "WTS": "white sucker", "WHCS": "white sucker", "CAP": "common carp",
  "GAR": "gar", "BOW": "bowfin (dogfish)", "GSF": "green sunfish",
  "PMK": "pumpkinseed", "HSF": "hybrid sunfish", "RBS": "rainbow smelt",
  "GOS": "golden shiner", "SPS": "spottail shiner", "FHM": "fathead minnow",
  "BGB": "bigmouth buffalo", "SHR": "shorthead redhorse", "GDF": "bowfin (dogfish)",
  "WHP": "white perch", "RKB": "rock bass", "BLK": "black buffalo",
  "QBK": "quillback", "SHB": "smallmouth buffalo", "GZS": "gizzard shad",
  "SCU": "sculpin", "IOD": "iowa darter", "JOD": "johnny darter",
  "OSS": "orangespotted sunfish", "TMD": "tadpole madtom", "BNM": "bluntnose minnow",
  "CMM": "central mudminnow", "SFS": "spotfin shiner", "SPO": "spottail shiner",
  "TPM": "tadpole madtom", "CNM": "central mudminnow", "BIB": "bigmouth buffalo",
  "MUE": "muskellunge", "EMS": "emerald shiner", "BOF": "bowfin (dogfish)",
  "BKF": "banded killifish", "BKS": "brook silverside", "BNS": "blacknose shiner",
  "JND": "johnny darter", "LGP": "logperch", "MMS": "mimic shiner",
  "RHS": "redhorse (misc)", "SHI": "shiner (misc)", "BST": "brook stickleback",
  "BLH": "bullhead (misc)", "SAB": "smallmouth buffalo", "OTS": "sucker (misc)",
  "SLR": "silver redhorse", "TME": "tiger muskellunge", "CRC": "creek chub",
  "GOF": "goldfish", "OTM": "minnow (misc)", "FCF": "flathead catfish",
  "LED": "least darter", "GIS": "gizzard shad", "SPT": "splake",
  "TLC": "tullibee (cisco)", "CSH": "common shiner", "BUB": "burbot",
  "LNS": "longnose sucker", "RCS": "river carpsucker", "WHB": "white bass",
  "SNG": "shortnose gar",
}

# Load user-provided overrides from a JSON file if it exists
_species_map_file = os.path.join(os.path.dirname(__file__), 'species_map.json')

def _load_species_map_file():
    """Reload species map from the JSON file on disk."""
    if os.path.exists(_species_map_file):
        try:
            with open(_species_map_file) as f:
                user_map = json.load(f)
                if isinstance(user_map, list):
                    for entry in user_map:
                        abbrev = entry.get('abbreviation', '').strip()
                        name = entry.get('species', '').strip()
                        if abbrev and name:
                            SPECIES_MAP[abbrev] = name.lower()
                elif isinstance(user_map, dict):
                    for k, v in user_map.items():
                        if k.strip() and str(v).strip():
                            SPECIES_MAP[k.strip()] = str(v).strip().lower()
        except Exception as e:
            print(f"Warning: could not load species_map.json: {e}")

_load_species_map_file()


def _species_name(abbrev):
    """Translate a species abbreviation to its common name."""
    return SPECIES_MAP.get(abbrev, abbrev)


def fetch_survey_data(lake_id):
    """
    Fetch lake survey data from the DNR JSON API.
    Returns all surveys as a list of record dicts.
    Each record is tagged with survey_type ('Standard Survey', 'Targeted Survey', etc.)
    """
    # Reload user overrides in case the file was updated
    _load_species_map_file()

    try:
        resp = requests.get(
            DNR_SURVEY_API_URL,
            params={'type': 'lake_survey', 'id': lake_id},
            timeout=30,
            headers={'User-Agent': 'MN-Fishing-Scout/1.0 (personal fishing tool)'}
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"Failed to fetch survey data for {lake_id}: {e}")
        return []

    result = data.get('result', {})
    surveys = result.get('surveys', [])
    if not surveys:
        return []

    records = []
    for survey in surveys:
        survey_date = survey.get('surveyDate', '')
        survey_year = _to_int(survey_date[:4]) if survey_date else None
        survey_type = survey.get('surveyType', '')
        survey_subtype = survey.get('surveySubType', '')

        for fish in survey.get('fishCatchSummaries', []):
            species_abbrev = fish.get('species', '')
            species_name = _species_name(species_abbrev)

            nr_cpue = fish.get('quartileCount')
            if nr_cpue == 'N/A':
                nr_cpue = None
            nr_weight = fish.get('quartileWeight')
            if nr_weight == 'N/A':
                nr_weight = None

            records.append({
                'lake_id': lake_id,
                'lake_name': '',
                'survey_year': survey_year,
                'survey_date': survey_date,
                'survey_type': survey_type,
                'survey_subtype': survey_subtype,
                'species': species_name,
                'species_abbrev': species_abbrev,
                'gear': fish.get('gear', ''),
                'cpue': _to_float(fish.get('CPUE')),
                'normal_range_cpue': nr_cpue,
                'avg_weight': _to_float(fish.get('averageWeight')),
                'normal_range_weight': nr_weight,
                'count': _to_int(fish.get('totalCatch')),
                'source_sheet': 'DNR Survey API',
            })

    return records


# --- Routes ---
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/lakes/add', methods=['POST'])
def add_lake():
    """Add a lake by DOW number: fetch metadata + scrape survey data from DNR."""
    try:
        data = request.get_json(force=True) or {}
    except Exception:
        data = {}
    lake_id = data.get('lake_id', '').strip()
    if not lake_id:
        return jsonify({'error': 'No lake ID provided'}), 400

    # Pad to 8 digits if needed
    lake_id = lake_id.replace('-', '')
    if len(lake_id) < 8:
        lake_id = lake_id.zfill(8)

    db = get_db()
    now = datetime.now().isoformat()
    errors = []

    # Step 1: Fetch metadata from DNR API
    try:
        meta = fetch_lake_metadata(lake_id)
    except Exception as e:
        meta = None
        errors.append(f'Metadata fetch error: {e}')

    lake_name = lake_id
    if meta:
        try:
            save_lake_metadata(db, lake_id, meta)
            lake_name = meta.get('name', lake_id)
        except Exception as e:
            errors.append(f'Error saving metadata: {e}')
    else:
        # Still create a minimal lake entry
        db.execute('''
            INSERT OR IGNORE INTO lakes (id, name, last_updated)
            VALUES (?, ?, ?)
        ''', (lake_id, lake_id, now))
        db.commit()
        if not errors:
            errors.append('Could not fetch lake metadata from DNR API (lake may still exist).')

    # Step 2: Fetch survey data from DNR JSON API
    try:
        survey_records = fetch_survey_data(lake_id)
    except Exception as e:
        survey_records = []
        errors.append(f'Survey fetch error: {e}')

    inserted = 0
    if survey_records:
        # Track unknown species abbreviations
        unknown_abbrevs = set()
        for r in survey_records:
            r['lake_name'] = lake_name
            if r.get('species_abbrev') and r['species'] == r['species_abbrev']:
                unknown_abbrevs.add(r['species_abbrev'])
            db.execute('''
                INSERT INTO survey_data (lake_id, lake_name, survey_year, survey_date,
                                         survey_type, species, gear, cpue,
                                         normal_range_cpue, avg_weight, normal_range_weight,
                                         count, source_sheet, uploaded_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (r['lake_id'], r['lake_name'], r['survey_year'], r.get('survey_date'),
                  r.get('survey_type', 'Standard Survey'), r['species'], r['gear'],
                  r['cpue'], r['normal_range_cpue'], r['avg_weight'], r['normal_range_weight'],
                  r['count'], r['source_sheet'], now))
            inserted += 1
        db.commit()
        if unknown_abbrevs:
            errors.append(
                f'Unknown species abbreviations (shown as-is): {", ".join(sorted(unknown_abbrevs))}. '
                'You can add translations in species_map.json in the app directory.'
            )
    else:
        if not errors:
            errors.append(
                'No survey data found for this lake from the DNR API. '
                'You can still manually upload CPUE data via Excel.'
            )

    survey_types = sorted(set(r.get('survey_type', '') for r in survey_records)) if survey_records else []

    return jsonify({
        'success': True,
        'lake_id': lake_id,
        'lake_name': lake_name,
        'metadata_found': meta is not None,
        'survey_records_imported': inserted,
        'survey_years': sorted(set(r['survey_year'] for r in survey_records if r.get('survey_year'))) if survey_records else [],
        'survey_types': survey_types,
        'warnings': errors
    })


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx/.xls files are supported'}), 400

    filename = secure_filename(f.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)

    records = parse_cpue_excel(filepath)
    if not records:
        return jsonify({'error': 'No CPUE data found in file. Check the format.'}), 400

    db = get_db()
    now = datetime.now().isoformat()
    inserted = 0
    lake_ids = set()
    for r in records:
        db.execute('''
            INSERT INTO survey_data (lake_id, lake_name, survey_year, species, gear, cpue,
                                     normal_range_cpue, avg_weight, normal_range_weight,
                                     count, source_sheet, uploaded_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (r['lake_id'], r['lake_name'], r['survey_year'], r['species'], r['gear'],
              r['cpue'], r['normal_range_cpue'], r['avg_weight'], r['normal_range_weight'],
              r['count'], r['source_sheet'], now))
        inserted += 1
        lake_ids.add(r['lake_id'])

    # Ensure lakes table has entries (basic ones from survey data)
    for lid in lake_ids:
        name = next((r['lake_name'] for r in records if r['lake_id'] == lid), None)
        db.execute('''
            INSERT OR IGNORE INTO lakes (id, name, last_updated)
            VALUES (?, ?, ?)
        ''', (lid, name, now))
    db.commit()

    return jsonify({
        'success': True,
        'records_inserted': inserted,
        'lakes_found': len(lake_ids),
        'lake_ids': list(lake_ids)
    })


@app.route('/api/lakes')
def api_lakes():
    db = get_db()
    lakes = db.execute('''
        SELECT l.*,
               (SELECT MAX(survey_year) FROM survey_data WHERE lake_id = l.id) as latest_survey,
               (SELECT COUNT(DISTINCT species) FROM survey_data WHERE lake_id = l.id) as species_count,
               (SELECT COUNT(*) FROM survey_data WHERE lake_id = l.id) as total_records
        FROM lakes l
        ORDER BY l.name
    ''').fetchall()
    return jsonify([dict(r) for r in lakes])


@app.route('/api/lakes/<lake_id>/refresh', methods=['POST'])
def refresh_lake(lake_id):
    try:
        meta = fetch_lake_metadata(lake_id)
        if meta:
            db = get_db()
            save_lake_metadata(db, lake_id, meta)
            return jsonify({'success': True, 'lake_name': meta.get('name', lake_id)})
        return jsonify({'error': 'Could not fetch metadata from DNR API'}), 502
    except Exception as e:
        return jsonify({'error': f'Error refreshing lake: {e}'}), 500


@app.route('/api/lakes/<lake_id>', methods=['DELETE'])
def delete_lake(lake_id):
    try:
        db = get_db()
        cur = db.execute('DELETE FROM survey_data WHERE lake_id = ?', (lake_id,))
        surveys_deleted = cur.rowcount
        cur = db.execute('DELETE FROM lakes WHERE id = ?', (lake_id,))
        lake_deleted = cur.rowcount > 0
        db.commit()
        return jsonify({
            'success': True,
            'lake_deleted': lake_deleted,
            'surveys_deleted': surveys_deleted
        })
    except Exception as e:
        return jsonify({'error': f'Error deleting lake: {e}'}), 500


@app.route('/api/lakes/<lake_id>')
def api_lake_detail(lake_id):
    db = get_db()
    lake = db.execute('SELECT * FROM lakes WHERE id = ?', (lake_id,)).fetchone()
    if not lake:
        return jsonify({'error': 'Lake not found'}), 404

    surveys = db.execute('''
        SELECT * FROM survey_data WHERE lake_id = ?
        ORDER BY survey_year DESC, species, gear
    ''', (lake_id,)).fetchall()

    return jsonify({
        'lake': dict(lake),
        'surveys': [dict(s) for s in surveys]
    })


import math

@app.route('/api/search')
def api_search():
    db = get_db()

    species = request.args.get('species', '').strip().lower()
    gear = request.args.get('gear', '').strip().lower()
    min_cpue = _to_float(request.args.get('min_cpue'))
    max_cpue = _to_float(request.args.get('max_cpue'))
    min_weight = _to_float(request.args.get('min_weight'))
    max_weight = _to_float(request.args.get('max_weight'))
    min_year = _to_int(request.args.get('min_year'))
    max_year = _to_int(request.args.get('max_year'))
    min_area = _to_float(request.args.get('min_area'))
    max_area = _to_float(request.args.get('max_area'))
    min_depth = _to_float(request.args.get('min_depth'))
    max_depth = _to_float(request.args.get('max_depth'))
    lake_name = request.args.get('lake_name', '').strip().lower()
    above_normal = request.args.get('above_normal', '').strip().lower() == 'true'
    county = request.args.get('county', '').strip()
    center_lat = _to_float(request.args.get('lat'))
    center_lng = _to_float(request.args.get('lng'))
    radius_miles = _to_float(request.args.get('radius'))
    survey_type = request.args.get('survey_type', 'standard').strip().lower()
    most_recent_only = request.args.get('most_recent', '').strip().lower() != 'false'

    # Default: exclude surveys older than 20 years
    current_year = datetime.now().year
    if min_year is None:
        min_year = current_year - 20

    conditions = ['s.survey_year >= ?']
    params = [min_year]

    # Default: only Standard Survey (unless 'all' is specified)
    if survey_type == 'standard':
        conditions.append("(s.survey_type = 'Standard Survey' OR s.survey_type IS NULL)")
    elif survey_type != 'all':
        conditions.append('LOWER(s.survey_type) LIKE ?')
        params.append(f'%{survey_type}%')

    if max_year:
        conditions.append('s.survey_year <= ?')
        params.append(max_year)
    if species:
        conditions.append('LOWER(s.species) LIKE ?')
        params.append(f'%{species}%')
    if gear:
        conditions.append('LOWER(s.gear) LIKE ?')
        params.append(f'%{gear}%')
    if min_cpue is not None:
        conditions.append('s.cpue >= ?')
        params.append(min_cpue)
    if max_cpue is not None:
        conditions.append('s.cpue <= ?')
        params.append(max_cpue)
    if min_weight is not None:
        conditions.append('s.avg_weight >= ?')
        params.append(min_weight)
    if max_weight is not None:
        conditions.append('s.avg_weight <= ?')
        params.append(max_weight)
    if lake_name:
        conditions.append('LOWER(COALESCE(l.name, s.lake_name)) LIKE ?')
        params.append(f'%{lake_name}%')
    if min_area is not None:
        conditions.append('l.area >= ?')
        params.append(min_area)
    if max_area is not None:
        conditions.append('l.area <= ?')
        params.append(max_area)
    if min_depth is not None:
        conditions.append('l.max_depth >= ?')
        params.append(min_depth)
    if max_depth is not None:
        conditions.append('l.max_depth <= ?')
        params.append(max_depth)
    if county:
        conditions.append('LOWER(l.county) = ?')
        params.append(county.lower())

    # For radius search, require that lake has coordinates
    if center_lat is not None and center_lng is not None and radius_miles is not None:
        conditions.append('l.latitude IS NOT NULL')
        conditions.append('l.longitude IS NOT NULL')

    where = ' AND '.join(conditions)

    # If most_recent_only, use a subquery to get each lake's max survey_year
    # (within the already-applied survey_type and year filters)
    if most_recent_only:
        query = f'''
            SELECT s.*, COALESCE(l.name, s.lake_name) as display_name,
                   l.county, l.area, l.max_depth, l.nearest_town, l.latitude, l.longitude
            FROM survey_data s
            LEFT JOIN lakes l ON s.lake_id = l.id
            INNER JOIN (
                SELECT s2.lake_id, MAX(s2.survey_year) as max_year
                FROM survey_data s2
                LEFT JOIN lakes l2 ON s2.lake_id = l2.id
                WHERE {where.replace('s.', 's2.').replace('l.', 'l2.')}
                GROUP BY s2.lake_id
            ) latest ON s.lake_id = latest.lake_id AND s.survey_year = latest.max_year
            WHERE {where}
            ORDER BY s.lake_id ASC, COALESCE(l.name, s.lake_name) ASC, s.species ASC
            LIMIT 2000
        '''
        results = db.execute(query, params + params).fetchall()
    else:
        query = f'''
            SELECT s.*, COALESCE(l.name, s.lake_name) as display_name,
                   l.county, l.area, l.max_depth, l.nearest_town, l.latitude, l.longitude
            FROM survey_data s
            LEFT JOIN lakes l ON s.lake_id = l.id
            WHERE {where}
            ORDER BY s.lake_id ASC, COALESCE(l.name, s.lake_name) ASC, s.survey_year DESC, s.species ASC
            LIMIT 2000
        '''
        results = db.execute(query, params).fetchall()

    results = [dict(r) for r in results]

    # Post-filter: radius from coordinates (Haversine)
    if center_lat is not None and center_lng is not None and radius_miles is not None:
        filtered_by_dist = []
        for r in results:
            lat = r.get('latitude')
            lng = r.get('longitude')
            if lat is not None and lng is not None:
                dist = _haversine_miles(center_lat, center_lng, lat, lng)
                r['distance_miles'] = round(dist, 1)
                if dist <= radius_miles:
                    filtered_by_dist.append(r)
        results = filtered_by_dist

    # Post-filter: above normal CPUE
    if above_normal:
        filtered = []
        for r in results:
            nr = r.get('normal_range_cpue')
            if nr and '-' in str(nr):
                try:
                    high = float(str(nr).split('-')[1])
                    if r['cpue'] and r['cpue'] > high:
                        r['above_normal'] = True
                        filtered.append(r)
                except (ValueError, IndexError):
                    pass
            else:
                filtered.append(r)
        results = filtered

    # Sort by distance if radius search, else by CPUE
    if center_lat is not None and center_lng is not None and radius_miles is not None:
        results.sort(key=lambda r: r.get('distance_miles', 9999))

    return jsonify(results[:500])


def _haversine_miles(lat1, lon1, lat2, lon2):
    R = 3958.8  # Earth radius in miles
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


@app.route('/api/species')
def api_species():
    db = get_db()
    species = db.execute('''
        SELECT DISTINCT species FROM survey_data ORDER BY species
    ''').fetchall()
    return jsonify([r['species'] for r in species])


@app.route('/api/gears')
def api_gears():
    db = get_db()
    gears = db.execute('''
        SELECT DISTINCT gear FROM survey_data ORDER BY gear
    ''').fetchall()
    return jsonify([r['gear'] for r in gears])


@app.route('/api/counties')
def api_counties():
    db = get_db()
    counties = db.execute('''
        SELECT DISTINCT county FROM lakes WHERE county IS NOT NULL AND county != '' ORDER BY county
    ''').fetchall()
    return jsonify([r['county'] for r in counties])


@app.route('/api/species_map')
def api_species_map():
    """Return the current species abbreviation mapping."""
    return jsonify(SPECIES_MAP)


@app.route('/api/species_map', methods=['POST'])
def update_species_map():
    """Add or update species abbreviation mappings. Persists to species_map.json and retranslates DB records."""
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    new_mappings = {}
    if isinstance(data, list):
        for entry in data:
            abbrev = entry.get('abbreviation', '').strip()
            name = entry.get('species', '').strip()
            if abbrev and name:
                SPECIES_MAP[abbrev] = name.lower()
                new_mappings[abbrev] = name.lower()
    elif isinstance(data, dict):
        for k, v in data.items():
            if k.strip() and str(v).strip():
                SPECIES_MAP[k.strip()] = str(v).strip().lower()
                new_mappings[k.strip()] = str(v).strip().lower()

    # Save to file for persistence
    try:
        with open(_species_map_file, 'w') as f:
            json.dump(SPECIES_MAP, f, indent=2)
    except Exception as e:
        return jsonify({'error': f'Saved in memory but failed to write file: {e}'}), 500

    # Retranslate existing DB records that match the updated abbreviations
    db = get_db()
    updated_rows = 0
    for abbrev, name in new_mappings.items():
        # Update records where species is the raw abbreviation
        cur = db.execute('UPDATE survey_data SET species = ? WHERE species = ?', (name, abbrev))
        updated_rows += cur.rowcount
    # Also retranslate ALL abbreviation-looking species names (2-4 uppercase chars)
    rows = db.execute("SELECT DISTINCT species FROM survey_data WHERE species = UPPER(species) AND LENGTH(species) <= 5").fetchall()
    for row in rows:
        sp = row['species']
        if sp in SPECIES_MAP:
            cur = db.execute('UPDATE survey_data SET species = ? WHERE species = ?', (SPECIES_MAP[sp], sp))
            updated_rows += cur.rowcount
    db.commit()

    return jsonify({'success': True, 'total_mappings': len(SPECIES_MAP), 'db_records_updated': updated_rows})


@app.route('/api/survey_types')
def api_survey_types():
    db = get_db()
    types = db.execute('''
        SELECT DISTINCT survey_type FROM survey_data
        WHERE survey_type IS NOT NULL AND survey_type != ''
        ORDER BY survey_type
    ''').fetchall()
    return jsonify([r['survey_type'] for r in types])


if __name__ == '__main__':
    app.run(debug=True, port=7002)
